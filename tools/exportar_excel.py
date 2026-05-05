"""
exportar_excel.py
Lê o arquivo Dispan_Banco_Completo_2023.xlsx e gera um arquivo de texto
com o conteúdo de todas as abas, pronto para colar no chat.
"""
import os
import sys

ARQUIVO = r"C:\Users\eliva\OneDrive\Área de Trabalho\Dispan_Banco_Completo_2023.xlsx"
SAIDA   = r"C:\Users\eliva\OneDrive\Área de Trabalho\Dispan_export.txt"

try:
    import openpyxl
except ImportError:
    print("openpyxl não encontrado. Rodando: pip install openpyxl")
    os.system("pip install openpyxl")
    import openpyxl

print(f"Abrindo: {ARQUIVO}")
wb = openpyxl.load_workbook(ARQUIVO, data_only=True)

linhas = []
for nome_aba in wb.sheetnames:
    ws = wb[nome_aba]
    linhas.append(f"\n{'='*60}")
    linhas.append(f"ABA: {nome_aba}")
    linhas.append(f"{'='*60}")

    # Detecta até onde há dados
    max_col = ws.max_column
    max_row = ws.max_row

    # Pula abas completamente vazias
    tem_dados = any(
        ws.cell(row=r, column=c).value is not None
        for r in range(1, min(max_row + 1, 5))
        for c in range(1, min(max_col + 1, 5))
    )
    if not tem_dados:
        linhas.append("(aba vazia)")
        continue

    linhas.append(f"Colunas: {max_col} | Linhas: {max_row}")
    linhas.append("")

    for row in ws.iter_rows(min_row=1, max_row=max_row, values_only=True):
        # Ignora linhas completamente vazias
        if all(v is None for v in row):
            continue
        celulas = [str(v).strip() if v is not None else "" for v in row]
        linhas.append("\t".join(celulas))

texto = "\n".join(linhas)

with open(SAIDA, "w", encoding="utf-8") as f:
    f.write(texto)

print(f"\nConcluído! Arquivo gerado em:\n{SAIDA}")
print(f"\nTotal de abas processadas: {len(wb.sheetnames)}")
print("Abas encontradas:", ", ".join(wb.sheetnames))
print("\nAbra o arquivo Dispan_export.txt e cole o conteúdo no chat.")
